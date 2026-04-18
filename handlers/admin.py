from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from telegram import Update
from telegram.ext import ContextTypes

from config import ADMIN_IDS
from database import (
    add_addon_item,
    add_menu_item,
    get_room_bookings_by_date,
    get_room_statuses,
    get_today_booking_stats,
    get_today_bookings,
    get_today_delivery_stats,
    get_today_free_rooms,
    get_delivery_orders_by_date,
    list_addons,
    list_menu_items,
    update_addon_price,
    update_booking_status,
    update_menu_item_price,
    update_order_status,
    conn,
)
from keyboards.inline import (
    addon_edit_keyboard,
    admin_order_keyboard,
    admin_panel_keyboard,
    date_keyboard,
    export_menu_keyboard,
    export_month_picker_keyboard,
    menu_edit_keyboard,
    menu_item_options_keyboard,
)
from keyboards.reply import main_menu_keyboard
from utils import format_price

ORDER_STATUS_BLOCK_TITLE = "📌 Buyurtma holati"


def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


def _admin_home_text() -> str:
    today = date.today().isoformat()
    free_rooms = get_today_free_rooms(today)
    bookings = get_today_bookings(today)
    delivery_totals, payment_rows, top_foods = get_today_delivery_stats(today)
    payment_map = {r["payment_method"]: r["cnt"] for r in payment_rows}

    lines = [
        "🛠 ADMIN PANEL",
        "",
        f"📅 Sana: {today}",
        "",
        "📌 TEZKOR DASHBOARD",
        f"• Buyurtmalar: {delivery_totals['total_orders']}",
        f"• Tushum: {format_price(delivery_totals['total_revenue'])}",
        f"• Naqd: {payment_map.get('cash', 0)} | Click: {payment_map.get('click', 0)}",
        f"• Bugungi bandlar: {len(bookings)}",
        f"• Bo'sh xonalar: {len(free_rooms)}",
        "",
        "🔥 TOP TAOMLAR",
    ]
    if top_foods:
        for i, row in enumerate(top_foods[:5], start=1):
            lines.append(f"{i}. {row['food_title']} — {row['cnt']} ta")
    else:
        lines.append("Hozircha buyurtmalar yo'q.")

    lines += [
        "",
        "Kerakli bo'limni tanlang:",
    ]
    return "\n".join(lines)


async def _edit_or_send(q, text: str, reply_markup=None):
    try:
        await q.edit_message_text(text=text, reply_markup=reply_markup)
    except Exception:
        await q.message.reply_text(text, reply_markup=reply_markup)


def build_stats_text(selected_date: str) -> str:
    delivery_totals, payment_rows, top_foods = get_today_delivery_stats(selected_date)
    booking_totals, booking_status_rows, booking_people = get_today_booking_stats(selected_date)
    payment_map = {r["payment_method"]: r["cnt"] for r in payment_rows}
    booking_map = {r["status"]: r["cnt"] for r in booking_status_rows}

    lines = [
        f"📊 {selected_date} statistikasi",
        "",
        "🛵 Yetkazib berish:",
        f"- Jami buyurtma: {delivery_totals['total_orders']}",
        f"- Umumiy tushum: {format_price(delivery_totals['total_revenue'])}",
        f"- Naqd: {payment_map.get('cash', 0)}",
        f"- Click: {payment_map.get('click', 0)}",
        "",
        "🔥 Eng ko'p buyurtma qilingan taomlar:",
    ]
    if top_foods:
        for i, row in enumerate(top_foods, start=1):
            lines.append(f"{i}. {row['food_title']} — {row['cnt']} ta")
    else:
        lines.append("Hozircha ma'lumot yo'q.")

    lines += [
        "",
        "🏠 Xona band qilish:",
        f"- Jami bandlar: {booking_totals['total_bookings']}",
        f"- Jami odam soni: {booking_people['total_people']}",
        f"- Yangi: {booking_map.get('new', 0)}",
        f"- Tasdiqlangan: {booking_map.get('confirmed', 0)}",
        f"- Bekor qilingan: {booking_map.get('cancelled', 0)}",
    ]
    return "\n".join(lines)


def build_menu_text() -> str:
    lines = ["🍽 Menu ro'yxati:"]
    for item in list_menu_items():
        option_text = " / ".join(
            f"{opt['option_label']} - {format_price(opt['price'])}" for opt in item["options"]
        )
        lines.append(f"• {item['title']}: {option_text}")
    return "\n".join(lines)


def build_addon_text() -> str:
    lines = ["🥗 Qo'shimchalar ro'yxati:"]
    for addon in list_addons():
        lines.append(f"• {addon['title']} — {format_price(addon['price'])}")
    return "\n".join(lines)

def _build_rooms_today_text(selected_date: str | None = None) -> str:
    selected_date = selected_date or date.today().isoformat()
    rows = get_room_statuses(selected_date)
    if not rows:
        return "Xonalar topilmadi."
    lines = [f"🏠 {selected_date} bo'yicha xonalar holati:\n"]
    for room in rows:
        if room["is_busy"]:
            lines.append(f"🔴 Xona {room['room_no']} ({room['capacity']} kishi) — BAND")
            for booking in room["bookings"]:
                lines.append(
                    f"   • {booking['booking_time']} | {booking['customer_name']} | {booking['customer_phone']} | {booking['people_count']} kishi"
                )
        else:
            lines.append(f"🟢 Xona {room['room_no']} ({room['capacity']} kishi) — BO'SH")
    return "\n".join(lines)



async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Siz admin emassiz.")
        return

    await update.message.reply_text(_admin_home_text(), reply_markup=admin_panel_keyboard())


async def free_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    today = date.today().isoformat()
    rows = get_today_free_rooms(today)
    if not rows:
        await update.message.reply_text("Bugun bo'sh xona yo'q.")
        return
    lines = [f"📋 {today} bo'yicha bo'sh xonalar:\n"]
    for r in rows:
        lines.append(f"Xona {r['room_no']} - {r['capacity']} kishilik")
    await update.message.reply_text("\n".join(lines))


async def bookings_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    today = date.today().isoformat()
    rows = get_today_bookings(today)
    if not rows:
        await update.message.reply_text("Bugun band xonalar yo'q.")
        return

    lines = [f"📌 {today} bugungi band xonalar:\n"]
    for r in rows:
        lines.append(
            f"#{r['id']} | {r['booking_time']} | Xona {r['room_no']} | "
            f"{r['people_count']} kishi | {r['customer_name']} | {r['status']} | {r['source']}"
        )
    await update.message.reply_text("\n".join(lines))


async def stats_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    await update.message.reply_text(build_stats_text(date.today().isoformat()))


async def rooms_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    await update.message.reply_text(
        _build_rooms_today_text(date.today().isoformat()),
        reply_markup=main_menu_keyboard(is_admin=True),
    )


def _append_or_replace_status_block(text: str, status_line: str) -> str:
    base_text = text or ""
    marker = f"\n\n{ORDER_STATUS_BLOCK_TITLE}\n"
    if marker in base_text:
        base_text = base_text.split(marker, 1)[0].rstrip()
    return f"{base_text}\n\n{ORDER_STATUS_BLOCK_TITLE}\n{status_line}"


async def _update_order_admin_message(
    context: ContextTypes.DEFAULT_TYPE,
    *,
    chat_id: int,
    message_id: int,
    base_text: str,
    order_id: int,
    user_id: int,
    status_line: str,
):
    new_text = _append_or_replace_status_block(base_text, status_line)
    keyboard = admin_order_keyboard(order_id, user_id)
    try:
        await context.bot.edit_message_text(
            chat_id=chat_id,
            message_id=message_id,
            text=new_text,
            reply_markup=keyboard,
        )
        return
    except Exception:
        pass

    try:
        await context.bot.edit_message_caption(
            chat_id=chat_id,
            message_id=message_id,
            caption=new_text,
            reply_markup=keyboard,
        )
    except Exception:
        pass


async def order_admin_action(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    _, action, order_id, user_id = q.data.split(":")
    order_id = int(order_id)
    user_id = int(user_id)
    base_text = (q.message.text or q.message.caption or "").split(
        f"\n\n{ORDER_STATUS_BLOCK_TITLE}\n", 1
    )[0].rstrip()

    if action == "cooking":
        context.user_data["pending_ready_time"] = {
            "order_id": order_id,
            "user_id": user_id,
            "message_chat_id": q.message.chat_id,
            "message_id": q.message.message_id,
            "base_text": base_text,
        }
        await q.message.reply_text(
            f"🕒 Buyurtma #{order_id} uchun taxminiy tayyor bo'lish vaqtini daqiqada yuboring.\nMasalan: 20"
        )
        return

    status_map = {"confirm": "confirmed", "sent": "sent", "cancel": "cancelled"}
    status = status_map[action]
    update_order_status(order_id, status)

    admin_status_map = {
        "confirmed": "✅ Qabul qilindi",
        "sent": "🚚 Jo'natildi",
        "cancelled": "❌ Bekor qilindi",
    }

    await _update_order_admin_message(
        context,
        chat_id=q.message.chat_id,
        message_id=q.message.message_id,
        base_text=base_text,
        order_id=order_id,
        user_id=user_id,
        status_line=admin_status_map[status],
    )

    msg_map = {
        "confirmed": "✅ Buyurtmangiz qabul qilindi.",
        "sent": "🚚 Buyurtmangiz kuryer tomonidan jo'natildi. Yoqimli ishtaha!",
        "cancelled": "❌ Sizning buyurtmangiz to'lov oxiriga yetkazilmaganligi sababli bekor qilindi.",
    }

    try:
        await context.bot.send_message(
            chat_id=user_id,
            text=f"{msg_map[status]}\nBuyurtma raqami: #{order_id}",
        )
    except Exception:
        pass


async def admin_ready_time_received(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message or not is_admin(update.effective_user.id):
        return

    pending = context.user_data.get("pending_ready_time")
    if pending:
        text = (update.message.text or "").strip()
        if not text.isdigit() or int(text) <= 0:
            await update.message.reply_text(
                "Iltimos, vaqtni daqiqada kiriting. Masalan: 20"
            )
            return

        minutes = int(text)
        order_id = pending["order_id"]
        user_id = pending["user_id"]
        update_order_status(order_id, "cooking")

        try:
            await _update_order_admin_message(
                context,
                chat_id=pending["message_chat_id"],
                message_id=pending["message_id"],
                base_text=pending["base_text"],
                order_id=order_id,
                user_id=user_id,
                status_line=f"🔄 Tayyorlanmoqda ({minutes} daqiqa)",
            )
        except Exception:
            pass

        try:
            await context.bot.send_message(
                chat_id=user_id,
                text=(
                    f"🔄 Buyurtmangiz jarayonga olindi va tayyorlanmoqda.\n"
                    f"⏱ Taxminiy tayyor bo'lish vaqti: {minutes} daqiqa.\n"
                    f"Buyurtma raqami: #{order_id}"
                ),
            )
        except Exception:
            pass

        await update.message.reply_text(
            f"✅ Buyurtma #{order_id} uchun tayyorlanish vaqti {minutes} daqiqa qilib belgilandi."
        )
        context.user_data.pop("pending_ready_time", None)
        return

    pending_action = context.user_data.get("pending_admin_action")
    if not pending_action:
        return

    mode = pending_action.get("mode")
    text = (update.message.text or "").strip()

    if mode == "update_menu_price":
        if not text.isdigit():
            await update.message.reply_text("Narxni faqat raqamda kiriting.")
            return
        update_menu_item_price(
            pending_action["item_key"], pending_action["option_label"], int(text)
        )
        await update.message.reply_text(
            "✅ Ovqat narxi yangilandi.", reply_markup=admin_panel_keyboard()
        )

    elif mode == "update_addon_price":
        if not text.isdigit():
            await update.message.reply_text("Narxni faqat raqamda kiriting.")
            return
        update_addon_price(pending_action["addon_key"], int(text))
        await update.message.reply_text(
            "✅ Qo'shimcha narxi yangilandi.", reply_markup=admin_panel_keyboard()
        )

    elif mode == "add_menu_item":
        parts = [p.strip() for p in text.split("|")]
        if len(parts) < 2 or not parts[1].isdigit():
            await update.message.reply_text(
                "Format: Nomi | narx | variant (ixtiyoriy) | eslatma (ixtiyoriy)"
            )
            return
        title = parts[0]
        price = int(parts[1])
        option_label = parts[2] if len(parts) >= 3 and parts[2] else "1 porsiya"
        note = parts[3] if len(parts) >= 4 and parts[3] else None
        add_menu_item(title, price, option_label, note)
        await update.message.reply_text(
            "✅ Yangi ovqat menuga qo'shildi.", reply_markup=admin_panel_keyboard()
        )

    elif mode == "add_addon_item":
        parts = [p.strip() for p in text.split("|")]
        if len(parts) < 2 or not parts[1].isdigit():
            await update.message.reply_text("Format: Nomi | narx")
            return
        add_addon_item(parts[0], int(parts[1]))
        await update.message.reply_text(
            "✅ Yangi qo'shimcha qo'shildi.", reply_markup=admin_panel_keyboard()
        )

    context.user_data.pop("pending_admin_action", None)


async def booking_admin_action(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    _, action, booking_id, user_id = q.data.split(":")

    status = "confirmed" if action == "confirm" else "cancelled"
    update_booking_status(int(booking_id), status)

    await q.edit_message_reply_markup(reply_markup=None)
    await q.message.reply_text(f"Band #{booking_id} holati: {status}")

    try:
        await context.bot.send_message(
            chat_id=int(user_id),
            text=f"🏠 Band holati yangilandi: {status}\nBand raqami: #{booking_id}",
        )
    except Exception:
        pass




def _month_range(year: int, month: int):
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1)
    else:
        end = date(year, month + 1, 1)
    return start.isoformat(), end.isoformat()


def _get_month_delivery_stats(year: int, month: int):
    start, end = _month_range(year, month)
    c = conn(); cur = c.cursor()
    cur.execute(
        """
        SELECT COUNT(*) AS total_orders, COALESCE(SUM(total_price), 0) AS total_revenue
        FROM delivery_orders
        WHERE DATE(created_at) >= DATE(?) AND DATE(created_at) < DATE(?) AND status != 'cancelled'
        """,
        (start, end),
    )
    totals = cur.fetchone()
    cur.execute(
        """
        SELECT payment_method, COUNT(*) AS cnt
        FROM delivery_orders
        WHERE DATE(created_at) >= DATE(?) AND DATE(created_at) < DATE(?) AND status != 'cancelled'
        GROUP BY payment_method
        """,
        (start, end),
    )
    payments = cur.fetchall()
    cur.execute(
        """
        SELECT food_title, COUNT(*) AS cnt
        FROM delivery_orders
        WHERE DATE(created_at) >= DATE(?) AND DATE(created_at) < DATE(?) AND status != 'cancelled'
        GROUP BY food_title
        ORDER BY cnt DESC, food_title ASC
        LIMIT 10
        """,
        (start, end),
    )
    top_foods = cur.fetchall()
    c.close()
    return totals, payments, top_foods


def _get_month_booking_stats(year: int, month: int):
    start, end = _month_range(year, month)
    c = conn(); cur = c.cursor()
    cur.execute(
        """
        SELECT COUNT(*) AS total_bookings
        FROM room_bookings
        WHERE DATE(booking_date) >= DATE(?) AND DATE(booking_date) < DATE(?) AND status NOT IN ('cancelled', 'completed')
        """,
        (start, end),
    )
    totals = cur.fetchone()
    cur.execute(
        """
        SELECT status, COUNT(*) AS cnt
        FROM room_bookings
        WHERE DATE(booking_date) >= DATE(?) AND DATE(booking_date) < DATE(?) AND status NOT IN ('cancelled', 'completed')
        GROUP BY status
        """,
        (start, end),
    )
    statuses = cur.fetchall()
    cur.execute(
        """
        SELECT COALESCE(SUM(people_count), 0) AS total_people
        FROM room_bookings
        WHERE DATE(booking_date) >= DATE(?) AND DATE(booking_date) < DATE(?) AND status NOT IN ('cancelled', 'completed')
        """,
        (start, end),
    )
    people = cur.fetchone()
    c.close()
    return totals, statuses, people


def _get_delivery_orders_by_month(year: int, month: int):
    start, end = _month_range(year, month)
    c = conn(); cur = c.cursor()
    cur.execute(
        "SELECT * FROM delivery_orders WHERE DATE(created_at) >= DATE(?) AND DATE(created_at) < DATE(?) ORDER BY created_at ASC, id ASC",
        (start, end),
    )
    rows = [dict(r) for r in cur.fetchall()]
    c.close()
    return rows


def _get_room_bookings_by_month(year: int, month: int):
    start, end = _month_range(year, month)
    c = conn(); cur = c.cursor()
    cur.execute(
        "SELECT * FROM room_bookings WHERE DATE(booking_date) >= DATE(?) AND DATE(booking_date) < DATE(?) ORDER BY booking_date ASC, booking_time ASC, room_no ASC",
        (start, end),
    )
    rows = [dict(r) for r in cur.fetchall()]
    c.close()
    return rows

def _build_export_workbook(selected_date: str) -> BytesIO:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Hisobot"
    ws1.append(["Sana", selected_date])
    ws1.append([])
    ws1.append(["Ko'rsatkich", "Qiymat"])

    delivery_totals, payment_rows, _ = get_today_delivery_stats(selected_date)
    booking_totals, booking_status_rows, booking_people = get_today_booking_stats(selected_date)
    payment_map = {r["payment_method"]: r["cnt"] for r in payment_rows}
    booking_map = {r["status"]: r["cnt"] for r in booking_status_rows}
    metrics = [
        ("Jami buyurtma", delivery_totals["total_orders"]),
        ("Umumiy tushum", delivery_totals["total_revenue"]),
        ("Naqd", payment_map.get("cash", 0)),
        ("Click", payment_map.get("click", 0)),
        ("Jami bandlar", booking_totals["total_bookings"]),
        ("Jami odam", booking_people["total_people"]),
        ("Yangi band", booking_map.get("new", 0)),
        ("Tasdiqlangan band", booking_map.get("confirmed", 0)),
        ("Bekor qilingan band", booking_map.get("cancelled", 0)),
    ]
    for row in metrics:
        ws1.append(list(row))

    ws2 = wb.create_sheet("Yetkazib berish")
    ws2.append(["ID", "Mijoz", "Telefon", "Taom", "Miqdor", "Qo'shimcha", "Jami", "To'lov", "Status", "Sana"])
    for row in get_delivery_orders_by_date(selected_date):
        ws2.append([
            row["id"],
            row["customer_name"],
            row["customer_phone"],
            row["food_title"],
            row["amount_value"],
            row["addons_json"],
            row["total_price"],
            row["payment_method"],
            row["status"],
            row["created_at"],
        ])

    ws3 = wb.create_sheet("Xona bandlar")
    ws3.append(["ID", "Mijoz", "Telefon", "Sana", "Vaqt", "Odam", "Xona", "Manba", "Status"])
    for row in get_room_bookings_by_date(selected_date):
        ws3.append([
            row["id"],
            row["customer_name"],
            row["customer_phone"],
            row["booking_date"],
            row["booking_time"],
            row["people_count"],
            row["room_no"],
            row["source"],
            row["status"],
        ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _build_month_export_workbook(year: int, month: int) -> BytesIO:
    month_label = f"{year}-{month:02d}"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Oylik hisobot"
    ws1.append(["Oy", month_label])
    ws1.append([])
    ws1.append(["Ko'rsatkich", "Qiymat"])

    delivery_totals, payment_rows, top_foods = _get_month_delivery_stats(year, month)
    booking_totals, booking_status_rows, booking_people = _get_month_booking_stats(year, month)
    payment_map = {r["payment_method"]: r["cnt"] for r in payment_rows}
    booking_map = {r["status"]: r["cnt"] for r in booking_status_rows}

    metrics = [
        ("Jami buyurtma", delivery_totals["total_orders"]),
        ("Umumiy tushum", delivery_totals["total_revenue"]),
        ("Naqd", payment_map.get("cash", 0)),
        ("Click", payment_map.get("click", 0)),
        ("Jami bandlar", booking_totals["total_bookings"]),
        ("Jami odam", booking_people["total_people"]),
        ("Yangi band", booking_map.get("new", 0)),
        ("Tasdiqlangan band", booking_map.get("confirmed", 0)),
        ("Bekor qilingan band", booking_map.get("cancelled", 0)),
    ]
    for row in metrics:
        ws1.append(list(row))

    ws1.append([])
    ws1.append(["Top taomlar", "Soni"])
    if top_foods:
        for row in top_foods:
            ws1.append([row["food_title"], row["cnt"]])
    else:
        ws1.append(["Ma'lumot yo'q", 0])

    ws2 = wb.create_sheet("Yetkazib berish")
    ws2.append(["ID", "Mijoz", "Telefon", "Taom", "Miqdor", "Qo'shimcha", "Jami", "To'lov", "Status", "Sana"])
    for row in _get_delivery_orders_by_month(year, month):
        ws2.append([
            row["id"], row["customer_name"], row["customer_phone"], row["food_title"],
            row["amount_value"], row["addons_json"], row["total_price"], row["payment_method"],
            row["status"], row["created_at"],
        ])

    ws3 = wb.create_sheet("Xona bandlar")
    ws3.append(["ID", "Mijoz", "Telefon", "Sana", "Vaqt", "Odam", "Xona", "Manba", "Status"])
    for row in _get_room_bookings_by_month(year, month):
        ws3.append([
            row["id"], row["customer_name"], row["customer_phone"], row["booking_date"],
            row["booking_time"], row["people_count"], row["room_no"], row["source"], row["status"],
        ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


async def admin_panel_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if not is_admin(q.from_user.id):
        await q.message.reply_text("Bu bo'lim faqat adminlar uchun.")
        return

    data = q.data
    if data == "admin:home":
        await _edit_or_send(q, _admin_home_text(), reply_markup=admin_panel_keyboard())
        return

    if data == "admin:rooms_today":
        today = date.today().isoformat()
        rows = get_room_statuses(today)
        if not rows:
            await _edit_or_send(q, "Xonalar topilmadi.", reply_markup=admin_panel_keyboard())
            return
        lines = [f"🏠 {today} bo'yicha xonalar holati:\n"]
        for room in rows:
            if room["is_busy"]:
                lines.append(f"🔴 Xona {room['room_no']} ({room['capacity']} kishi) — BAND")
                for booking in room["bookings"]:
                    lines.append(
                        f"   • {booking['booking_time']} | {booking['customer_name']} | "
                        f"{booking['customer_phone']} | {booking['people_count']} kishi"
                    )
            else:
                lines.append(f"🟢 Xona {room['room_no']} ({room['capacity']} kishi) — BO'SH")
        await _edit_or_send(q, "\n".join(lines), reply_markup=admin_panel_keyboard())
        return

    if data == "admin:add_booking":
        from handlers.booking import admin_add_booking_start
        return await admin_add_booking_start(update, context)

    if data == "admin:rooms_today":
        await _edit_or_send(q, _build_rooms_today_text(date.today().isoformat()), reply_markup=admin_panel_keyboard())
        return

    if data == "admin:report_date_picker":
        await _edit_or_send(
            q,
            "📅 Hisobot uchun kunni tanlang:",
            reply_markup=date_keyboard(prefix="admin_report_date", nav_prefix="admin_report_nav", back_callback='admin:home'),
        )
        return

    if data == "admin:export_menu":
        await _edit_or_send(
            q,
            "📤 Excel export turini tanlang:",
            reply_markup=export_menu_keyboard(),
        )
        return

    if data == "admin:export_today":
        today = date.today().isoformat()
        bio = _build_export_workbook(today)
        bio.name = f"hisobot_{today}.xlsx"
        await context.bot.send_document(
            chat_id=q.message.chat_id,
            document=bio,
            filename=bio.name,
            caption=f"📤 {today} bo'yicha Excel hisobot",
        )
        await q.answer("Excel yuborildi")
        return

    if data == "admin:export_month":
        today = date.today()
        await _edit_or_send(
            q,
            "📆 Excel export uchun oyni tanlang:",
            reply_markup=export_month_picker_keyboard(today.year),
        )
        return

    if data.startswith("admin:export_month_nav:"):
        parts = data.split(":")
        if len(parts) < 4:
            await q.answer("Yil ma'lumotida xato bor", show_alert=True)
            return
        year = int(parts[-1])
        await q.message.edit_reply_markup(reply_markup=export_month_picker_keyboard(year))
        return

    if data.startswith("admin:export_month_select:"):
        parts = data.split(":")
        if len(parts) != 5:
            await q.answer("Oylik export ma'lumotida xato bor", show_alert=True)
            return
        _, _, _, year, month = parts
        year = int(year)
        month = int(month)
        bio = _build_month_export_workbook(year, month)
        await context.bot.send_document(
            chat_id=q.message.chat_id,
            document=bio,
            filename=f"hisobot_{year}_{month:02d}.xlsx",
            caption=f"📤 {year}-{month:02d} oylik Excel hisobot",
        )
        await q.answer("Excel yuborildi")
        return

    if data == "admin:export_date_picker":
        await _edit_or_send(
            q,
            "📅 Excel export uchun kunni tanlang:",
            reply_markup=date_keyboard(prefix="admin_export_date", nav_prefix="admin_export_nav", back_callback='admin:export_menu'),
        )
        return

    if data.startswith("admin_report_nav:"):
        parts = data.split(":")
        if len(parts) != 3:
            await q.answer("Kalendar ma'lumotida xato bor", show_alert=True)
            return
        _, year, month = parts
        await q.message.edit_reply_markup(
            reply_markup=date_keyboard(int(year), int(month), prefix="admin_report_date", nav_prefix="admin_report_nav", back_callback='admin:home')
        )
        return

    if data.startswith("admin_export_nav:"):
        parts = data.split(":")
        if len(parts) != 3:
            await q.answer("Kalendar ma'lumotida xato bor", show_alert=True)
            return
        _, year, month = parts
        await q.message.edit_reply_markup(
            reply_markup=date_keyboard(int(year), int(month), prefix="admin_export_date", nav_prefix="admin_export_nav", back_callback='admin:export_menu')
        )
        return

    if data.startswith("admin_report_date:"):
        selected_date = data.split(":", 1)[1]
        await _edit_or_send(q, build_stats_text(selected_date), reply_markup=admin_panel_keyboard())
        return

    if data.startswith("admin_export_date:"):
        selected_date = data.split(":", 1)[1]
        bio = _build_export_workbook(selected_date)
        await context.bot.send_document(
            chat_id=q.message.chat_id,
            document=bio,
            filename=f"hisobot_{selected_date}.xlsx",
            caption=f"📤 {selected_date} sanasi bo'yicha Excel hisobot",
        )
        await q.answer("Excel yuborildi")
        return

    if data == "admin:menu_edit":
        await _edit_or_send(q, build_menu_text(), reply_markup=menu_edit_keyboard())
        return

    if data.startswith("admin:menu_item:"):
        item_key = data.split(":", 2)[2]
        await _edit_or_send(
            q,
            "Qaysi variant narxini o'zgartirasiz?",
            reply_markup=menu_item_options_keyboard(item_key),
        )
        return

    if data.startswith("admin:menu_price:"):
        _, _, item_key, option_label = data.split(":", 3)
        context.user_data["pending_admin_action"] = {
            "mode": "update_menu_price",
            "item_key": item_key,
            "option_label": option_label,
        }
        await _edit_or_send(q, f"Yangi narxni yuboring:\n{item_key} / {option_label}")
        return

    if data == "admin:addon_edit":
        await _edit_or_send(q, build_addon_text(), reply_markup=addon_edit_keyboard())
        return

    if data.startswith("admin:addon_price:"):
        addon_key = data.split(":", 2)[2]
        context.user_data["pending_admin_action"] = {
            "mode": "update_addon_price",
            "addon_key": addon_key,
        }
        await _edit_or_send(q, "Yangi qo'shimcha narxini yuboring:")
        return

    if data == "admin:menu_add":
        context.user_data["pending_admin_action"] = {"mode": "add_menu_item"}
        await _edit_or_send(
            q,
            "Yangi ovqat qo'shish formati:\nNomi | narx | variant (ixtiyoriy) | eslatma (ixtiyoriy)",
        )
        return

    if data == "admin:addon_add":
        context.user_data["pending_admin_action"] = {"mode": "add_addon_item"}
        await _edit_or_send(q, "Yangi qo'shimcha formati:\nNomi | narx")
        return
